from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import sys
import json
import configparser
from pathlib import Path
import shutil
import re
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLabel,
    QTextEdit,
    QComboBox,
    QFileDialog,
    QMessageBox,
    QLineEdit,
    QTabWidget,
    QCheckBox,
    QGroupBox,
)
from PySide6.QtCore import Qt, QTimer, Signal, QObject
from datetime import datetime
from openpyxl import Workbook, load_workbook

from sms_sender import send_sms
from crawler import SiteCrawler

CONFIG_PATH = Path(__file__).with_name("config.ini")
TEMPLATE_PATH = Path(__file__).with_name("templates.json")
LOG_XLSX_PATH = Path(__file__).with_name("sms_logs.xlsx")


class CrawlerTestWorker(QObject):
    """크롤러 테스트를 위한 워커 클래스"""
    status_updated = Signal(str)  # 상태 업데이트 시그널
    
    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
    
    def run_test(self):
        """크롤러 테스트 실행 (전체 스캔 방식)"""
        try:

            
            url = self.cfg.get("web", "url")
            username = self.cfg.get("web", "username")
            password = self.cfg.get("web", "password")
            
            options = Options()
            options.add_argument("--headless=new")
            driver = webdriver.Chrome(options=options)
            
            self.status_updated.emit(f"웹사이트 접속: {url}")
            driver.get(url)
            
            # 로그인
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="id"]'))
            )
            
            driver.find_element(By.XPATH, '//*[@id="id"]').send_keys(username)
            driver.find_element(By.XPATH, '//*[@id="pwd"]').send_keys(password)
            driver.find_element(By.XPATH, '//*[@id="pwd"]').submit()
            
            self.status_updated.emit("로그인 완료")
            
            # 테이블 확인
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/section/div/div[2]/table/tbody/tr[1]/td[1]'))
            )
            
            # 전체 테이블 스캔 테스트
            consultations = []
            self.status_updated.emit("📊 전체 테이블 스캔 시작...")
            
            for row_idx in range(1, 11):  # 상위 10개 행만 테스트
                try:
                    xpath_no = f'/html/body/section/div/div[2]/table/tbody/tr[{row_idx}]/td[1]'
                    xpath_phone = f'/html/body/section/div/div[2]/table/tbody/tr[{row_idx}]/td[16]'
                    
                    no_element = driver.find_elements(By.XPATH, xpath_no)
                    if not no_element:
                        break
                    
                    consultation_no = no_element[0].text.strip()
                    if not consultation_no:
                        break
                    
                    phone_element = driver.find_elements(By.XPATH, xpath_phone)
                    phone = phone_element[0].text.strip() if phone_element else ""
                    
                    if consultation_no and phone:
                        consultations.append({"no": consultation_no, "phone": phone})
                        self.status_updated.emit(f"행 {row_idx}: No.{consultation_no} - {phone}")
                    
                except Exception as e:
                    self.status_updated.emit(f"행 {row_idx} 처리 중 오류: {e}")
                    break
            
            self.status_updated.emit(f"✅ 크롤러 테스트 성공! 총 {len(consultations)}건 확인")
            
            # 첫 번째와 마지막 상담 정보 출력
            if consultations:
                first = consultations[0]
                last = consultations[-1]
                self.status_updated.emit(f"📋 첫 번째: No.{first['no']} - {first['phone']}")
                self.status_updated.emit(f"📋 마지막: No.{last['no']} - {last['phone']}")
            
            driver.quit()
            
        except Exception as e:
            self.status_updated.emit(f"❌ 크롤러 테스트 실패: {e}")


def load_templates():
    if TEMPLATE_PATH.exists():
        with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_templates(data):
    with open(TEMPLATE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

class TemplateManagerTab(QWidget):
    def __init__(self, sms_tab_ref=None):
        super().__init__()
        self.templates = load_templates()
        self.sms_tab_ref = sms_tab_ref  # SmsSenderTab 참조
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        self.template_list = QComboBox()
        self.template_list.addItems([t.get("name", "템플릿") for t in self.templates])
        self.template_list.currentIndexChanged.connect(self._on_select)

        self.name_edit = QLineEdit()
        self.content_edit = QTextEdit()
        
        # 실제 한국 SMS 화면처럼 스타일 적용
        phone_style = """
            QTextEdit {
                background-color: #e8f4fd;
                border: 2px solid #007acc;
                border-radius: 15px;
                padding: 12px 16px;
                font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif;
                font-size: 15px;
                line-height: 1.5;
                max-width: 280px;
                min-width: 260px;
                max-height: 300px;
            }
        """
        self.content_edit.setStyleSheet(phone_style)
        self.content_edit.setMaximumWidth(320)
        self.content_edit.setMaximumHeight(320)  # 편집용이므로 조금 더 높게

        button_layout = QHBoxLayout()
        new_btn = QPushButton("새 템플릿")
        new_btn.clicked.connect(self._on_new_template)
        
        save_btn = QPushButton("저장")
        save_btn.clicked.connect(self._on_save)
        
        delete_btn = QPushButton("삭제")
        delete_btn.clicked.connect(self._on_delete_template)
        
        button_layout.addWidget(new_btn)
        button_layout.addWidget(save_btn)
        button_layout.addWidget(delete_btn)

        layout.addWidget(QLabel("템플릿 목록"))
        layout.addWidget(self.template_list)
        layout.addWidget(QLabel("이름"))
        layout.addWidget(self.name_edit)
        
        # 템플릿 내용과 글자수 카운터
        content_container = QVBoxLayout()
        content_label_layout = QHBoxLayout()
        
        content_label = QLabel("내용")
        self.template_char_count_label = QLabel("0/90byte (SMS) 0글자")
        self.template_char_count_label.setStyleSheet("color: #666; font-size: 12px;")
        
        content_label_layout.addWidget(content_label)
        content_label_layout.addStretch()
        content_label_layout.addWidget(self.template_char_count_label)
        
        content_container.addLayout(content_label_layout)
        content_container.addWidget(self.content_edit)
        
        # 텍스트 변경 시 글자수 업데이트
        self.content_edit.textChanged.connect(self._update_template_char_count)
        
        layout.addLayout(content_container)
        layout.addLayout(button_layout)

        self._on_select(0)

    def _on_select(self, idx):
        if 0 <= idx < len(self.templates):
            t = self.templates[idx]
            self.name_edit.setText(t.get("name", ""))
            self.content_edit.setText(t.get("content", ""))
            self._update_template_char_count()
        else:
            self.name_edit.clear()
            self.content_edit.clear()
            self._update_template_char_count()
    
    def _calculate_byte_size(self, text):
        """한국 SMS 기준 바이트 계산 (EUC-KR 기준)"""
        try:
            # EUC-KR 인코딩으로 바이트 수 계산
            return len(text.encode('euc-kr'))
        except UnicodeEncodeError:
            # EUC-KR로 인코딩 불가능한 문자가 있으면 UTF-8로 대체 계산
            return len(text.encode('utf-8'))
    
    def _update_template_char_count(self):
        """템플릿 편집 시 바이트 기준 카운트"""
        content = self.content_edit.toPlainText()
        byte_count = self._calculate_byte_size(content)
        char_count = len(content)
        
        if byte_count <= 90:
            msg_type = "SMS"
            color = "#007acc"
            limit = "90"
        elif byte_count <= 2000:
            msg_type = "LMS"
            color = "#ff8c00"
            limit = "2,000"
        else:
            msg_type = "초과"
            color = "#ff4444"
            limit = "2,000"
        
        self.template_char_count_label.setText(f"{byte_count}/{limit}byte ({msg_type}) {char_count}글자")
        self.template_char_count_label.setStyleSheet(f"color: {color}; font-size: 12px; font-weight: bold;")

    def _on_new_template(self):
        """새 템플릿 추가"""
        self.templates.append({"name": "새 템플릿", "content": ""})
        self.template_list.addItem("새 템플릿")
        self.template_list.setCurrentIndex(len(self.templates) - 1)
        self._on_select(len(self.templates) - 1)
        self.name_edit.selectAll()  # 이름 전체 선택해서 바로 수정 가능
        self.name_edit.setFocus()
    
    def _on_delete_template(self):
        """템플릿 삭제"""
        idx = self.template_list.currentIndex()
        if idx < 0 or idx >= len(self.templates):
            QMessageBox.warning(self, "경고", "삭제할 템플릿을 선택해주세요.")
            return
        
        template_name = self.templates[idx].get("name", "템플릿")
        reply = QMessageBox.question(self, "삭제 확인", 
                                   f"'{template_name}' 템플릿을 삭제하시겠습니까?",
                                   QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            del self.templates[idx]
            self.template_list.removeItem(idx)
            save_templates(self.templates)
            
            # SMS 탭에서 템플릿 새로고침
            if self.sms_tab_ref:
                self.sms_tab_ref._refresh_templates()
            
            QMessageBox.information(self, "삭제", "템플릿이 삭제되었습니다.")
    
    def _on_save(self):
        name = self.name_edit.text().strip()
        content = self.content_edit.toPlainText().strip()
        
        if not name:
            QMessageBox.warning(self, "경고", "템플릿 이름을 입력해주세요.")
            return
        
        if not content:
            QMessageBox.warning(self, "경고", "템플릿 내용을 입력해주세요.")
            return
        
        idx = self.template_list.currentIndex()
        tmpl = {
            "name": name,
            "content": content,
        }
        
        if idx >= 0 and idx < len(self.templates):
            self.templates[idx] = tmpl
            self.template_list.setItemText(idx, tmpl["name"])
        else:
            self.templates.append(tmpl)
            self.template_list.addItem(tmpl["name"])
        
        save_templates(self.templates)
        
        # SMS 탭에서 템플릿 새로고침
        if self.sms_tab_ref:
            self.sms_tab_ref._refresh_templates()
        
        QMessageBox.information(self, "저장", "템플릿이 저장되었습니다.")


class LogViewerTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        
        # SMS 로그와 크롤러 로그를 탭으로 분리
        log_tabs = QTabWidget()
        
        # SMS 발송 로그 탭
        sms_log_widget = QWidget()
        sms_layout = QVBoxLayout(sms_log_widget)
        self.sms_text_view = QTextEdit()
        self.sms_text_view.setReadOnly(True)
        
        # SMS 로그 버튼 레이아웃
        sms_button_layout = QHBoxLayout()
        sms_refresh_btn = QPushButton("SMS 로그 새로고침")
        sms_refresh_btn.clicked.connect(self._load_sms_logs)
        sms_download_btn = QPushButton("📥 SMS 로그 엑셀 다운로드")
        sms_download_btn.clicked.connect(self._download_sms_logs)
        sms_button_layout.addWidget(sms_refresh_btn)
        sms_button_layout.addWidget(sms_download_btn)
        
        sms_layout.addWidget(self.sms_text_view)
        sms_layout.addLayout(sms_button_layout)
        
        # 크롤러 로그 탭
        crawler_log_widget = QWidget()
        crawler_layout = QVBoxLayout(crawler_log_widget)
        self.crawler_text_view = QTextEdit()
        self.crawler_text_view.setReadOnly(True)
        
        # 크롤러 로그 버튼 레이아웃
        crawler_button_layout = QHBoxLayout()
        crawler_refresh_btn = QPushButton("크롤러 로그 새로고침")
        crawler_refresh_btn.clicked.connect(self._load_crawler_logs)
        crawler_download_btn = QPushButton("📥 크롤러 로그 엑셀 다운로드")
        crawler_download_btn.clicked.connect(self._download_crawler_logs)
        crawler_button_layout.addWidget(crawler_refresh_btn)
        crawler_button_layout.addWidget(crawler_download_btn)
        
        crawler_layout.addWidget(self.crawler_text_view)
        crawler_layout.addLayout(crawler_button_layout)
        
        log_tabs.addTab(sms_log_widget, "SMS 발송 로그")
        log_tabs.addTab(crawler_log_widget, "크롤러 로그")
        
        layout.addWidget(log_tabs)
        
        self._load_sms_logs()
        self._load_crawler_logs()

    def _load_sms_logs(self):
        if not LOG_XLSX_PATH.exists():
            self.sms_text_view.setText("SMS 로그 파일이 없습니다.")
            return
        try:
            wb = load_workbook(LOG_XLSX_PATH, read_only=True)
            ws = wb.active
            rows = ["\t".join([str(c.value) for c in row]) for row in ws.iter_rows(values_only=True)]
            self.sms_text_view.setText("\n".join(rows))
        except Exception as e:
            self.sms_text_view.setText(f"SMS 로그 로드 오류: {e}")
    
    def _load_crawler_logs(self):
        crawler_log_path = Path(__file__).with_name("crawler.log")
        if not crawler_log_path.exists():
            self.crawler_text_view.setText("크롤러 로그 파일이 없습니다.\n크롤러가 아직 시작되지 않았거나 로그가 생성되지 않았습니다.")
            return
        try:
            with open(crawler_log_path, "r", encoding="utf-8") as f:
                content = f.read()
                # 최근 100줄만 표시
                lines = content.strip().split('\n')
                if len(lines) > 100:
                    lines = lines[-100:]
                    content = "... (이전 로그 생략) ...\n" + "\n".join(lines)
                self.crawler_text_view.setText(content)
        except Exception as e:
            self.crawler_text_view.setText(f"크롤러 로그 로드 오류: {e}")
    
    def _download_sms_logs(self):
        """SMS 로그를 엑셀 파일로 다운로드"""
        if not LOG_XLSX_PATH.exists():
            QMessageBox.warning(self, "다운로드 오류", "SMS 로그 파일이 존재하지 않습니다.\n먼저 SMS를 발송해주세요.")
            return
        
        # 파일 저장 대화상자
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"SMS_발송로그_{current_time}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "SMS 로그 엑셀 다운로드",
            default_filename,
            "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if file_path:
            try:
                # 원본 파일을 선택한 위치에 복사
                shutil.copy2(LOG_XLSX_PATH, file_path)
                QMessageBox.information(self, "다운로드 완료", f"SMS 로그가 성공적으로 다운로드되었습니다:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "다운로드 오류", f"파일 다운로드 중 오류가 발생했습니다:\n{e}")
    
    def _download_crawler_logs(self):
        """크롤러 로그를 엑셀 파일로 다운로드"""
        crawler_log_path = Path(__file__).with_name("crawler.log")
        if not crawler_log_path.exists():
            QMessageBox.warning(self, "다운로드 오류", "크롤러 로그 파일이 존재하지 않습니다.\n크롤러가 아직 시작되지 않았습니다.")
            return
        
        # 파일 저장 대화상자
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"크롤러_로그_{current_time}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "크롤러 로그 엑셀 다운로드",
            default_filename,
            "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if file_path:
            try:
                # 텍스트 로그를 엑셀로 변환
                wb = Workbook()
                ws = wb.active
                ws.title = "크롤러 로그"
                
                # 헤더 추가
                ws.append(["번호", "시간", "로그 내용"])
                
                # 로그 파일 읽기
                with open(crawler_log_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                
                # 각 로그 라인 파싱하여 엑셀에 추가
                for idx, line in enumerate(lines, 1):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # 로그 형식: YYYY-MM-DD HH:MM:SS,milliseconds - LEVEL - message
                    # 시간 패턴 추출 시도
                    time_pattern = r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ - (\w+) - (.+)$'
                    match = re.match(time_pattern, line)
                    
                    if match:
                        timestamp = match.group(1)  # 날짜+시간 부분
                        level = match.group(2)      # 로그 레벨 (INFO, ERROR 등)
                        message = match.group(3)    # 실제 메시지
                        ws.append([idx, timestamp, f"[{level}] {message}"])
                    else:
                        # 시간 패턴이 없는 경우 전체를 메시지로 처리
                        ws.append([idx, "", line])
                
                # 컬럼 너비 자동 조정
                ws.column_dimensions['A'].width = 8   # 번호
                ws.column_dimensions['B'].width = 20  # 시간
                ws.column_dimensions['C'].width = 80  # 로그 내용
                
                # 엑셀 파일 저장
                wb.save(file_path)
                QMessageBox.information(self, "다운로드 완료", f"크롤러 로그가 성공적으로 엑셀로 다운로드되었습니다:\n{file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "다운로드 오류", f"엑셀 파일 생성 중 오류가 발생했습니다:\n{e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("사이트 감지 & 문자발송")

        # Load config
        self.cfg = configparser.ConfigParser()
        self.cfg.read(CONFIG_PATH, encoding="utf-8")

        # Tabs
        self.sms_tab = SmsSenderTab(self.cfg)
        tab = QTabWidget()
        tab.addTab(self.sms_tab, "문자발송")
        tab.addTab(TemplateManagerTab(self.sms_tab), "템플릿관리")
        tab.addTab(LogViewerTab(), "발송내역확인")
        
        # 크롤러 디버깅 탭 추가
        debug_tab = QWidget()
        debug_layout = QVBoxLayout(debug_tab)
        
        # 크롤러 상태 그룹
        crawler_group = QGroupBox("크롤러 상태")
        crawler_group_layout = QVBoxLayout()
        
        status_label = QLabel("크롤러 동작 로그:")
        self.status_text = QTextEdit()
        self.status_text.setMaximumHeight(120)
        self.status_text.setReadOnly(True)
        
        test_crawl_btn = QPushButton("크롤러 수동 테스트")
        test_crawl_btn.clicked.connect(self._test_crawler)
        
        crawler_group_layout.addWidget(status_label)
        crawler_group_layout.addWidget(self.status_text)
        crawler_group_layout.addWidget(test_crawl_btn)
        crawler_group.setLayout(crawler_group_layout)
        
        # 발송 이력 정보 그룹
        history_group = QGroupBox("발송 이력 현황")
        history_layout = QVBoxLayout()
        
        self.history_info_text = QTextEdit()
        self.history_info_text.setMaximumHeight(100)
        self.history_info_text.setReadOnly(True)
        
        refresh_history_btn = QPushButton("발송 이력 새로고침")
        refresh_history_btn.clicked.connect(self._update_history_info)
        
        history_layout.addWidget(self.history_info_text)
        history_layout.addWidget(refresh_history_btn)
        history_group.setLayout(history_layout)
        
        # SENS 설정 확인 영역
        sens_group = QGroupBox("SENS SMS 설정 확인")
        sens_layout = QVBoxLayout()
        
        self.sens_info_text = QTextEdit()
        self.sens_info_text.setMaximumHeight(120)
        self.sens_info_text.setReadOnly(True)
        
        sens_test_btn = QPushButton("SMS 설정 테스트")
        sens_test_btn.clicked.connect(self._test_sms_config)
        
        sens_layout.addWidget(self.sens_info_text)
        sens_layout.addWidget(sens_test_btn)
        sens_group.setLayout(sens_layout)
        
        debug_layout.addWidget(crawler_group)
        debug_layout.addWidget(history_group)
        debug_layout.addWidget(sens_group)
        debug_layout.addStretch()
        
        # 초기 정보 표시
        self._update_sens_info()
        self._update_history_info()
        
        tab.addTab(debug_tab, "크롤러 디버그")

        self.setCentralWidget(tab)

        # Start crawler in background and 자동 문자발송 연동
        self.crawler = SiteCrawler(self.cfg, self._on_new_consultations, self._update_status)
        
        # 크롤러 테스트 워커 초기화
        self.test_worker = CrawlerTestWorker(self.cfg)
        self.test_worker.status_updated.connect(self._update_status)
        
        # 초기 시작 상태 표시
        self._show_startup_info()

    def _test_crawler(self):
        """크롤러 수동 테스트"""
        self.status_text.setText("크롤러 테스트 시작...")
        
        # QTimer를 사용해서 안전하게 백그라운드에서 실행
        QTimer.singleShot(100, self.test_worker.run_test)
    
    def _update_status(self, message):
        """상태 메시지 업데이트 (메인 스레드에서 안전하게 실행)"""
        self.status_text.append(message)
    
    def _update_sens_info(self):
        """SENS 설정 정보 표시"""
        sens_cfg = self.cfg["sens"]
        access_key = sens_cfg.get("access_key", fallback="")
        secret_key = sens_cfg.get("secret_key", fallback="")
        service_id = sens_cfg.get("service_id", fallback="")
        sender = sens_cfg.get("sender", fallback="")
        
        info_text = f"""SENS SMS 설정 현황:
• Access Key: {access_key[:8]}...{"*" * 8} (총 {len(access_key)}자)
• Secret Key: {secret_key[:8]}...{"*" * 8} (총 {len(secret_key)}자)
• Service ID: {service_id}
• 발신번호: {sender}

⚠️ config.ini 파일에서 실제 키 값을 확인하세요."""
        
        self.sens_info_text.setText(info_text)
    
    def _test_sms_config(self):
        """SMS 설정 테스트"""
        try:
            # 간단한 SMS 발송 테스트 (테스트 번호로)
            test_content = "SENS 설정 테스트 메시지"
            test_number = self.cfg.get("settings", "test_receiver", fallback="01012345678")
            
            self.sens_info_text.append(f"\n📱 테스트 발송 시작...")
            self.sens_info_text.append(f"수신번호: {test_number}")
            self.sens_info_text.append(f"내용: {test_content}")
            
            result = send_sms(test_number, test_content, self.cfg)
            
            if result:
                self.sens_info_text.append("✅ SMS 설정 테스트 성공!")
                QMessageBox.information(self, "테스트 성공", "SMS 설정이 올바르게 구성되었습니다.")
            else:
                self.sens_info_text.append("❌ SMS 설정 테스트 실패!")
                QMessageBox.critical(self, "테스트 실패", 
                                   "SMS 설정에 문제가 있습니다.\n"
                                   "콘솔 로그를 확인하거나 다음을 점검하세요:\n\n"
                                   "1. Access Key & Secret Key 확인\n"
                                   "2. Service ID 확인\n"
                                   "3. 발신번호 등록 상태 확인\n"
                                   "4. 네이버 클라우드 계정 권한 확인")
                
        except Exception as e:
            self.sens_info_text.append(f"❌ 테스트 중 오류 발생: {e}")
            QMessageBox.critical(self, "테스트 오류", f"테스트 중 오류가 발생했습니다:\n{e}")

    def _on_new_consultations(self, consultations):
        """크롤러에서 신규 상담들 감지 시 자동 발송 처리
        
        Args:
            consultations: [{"no": "3229", "phone": "01012345678"}, ...] 형태의 리스트
        """
        if not self.sms_tab.templates:
            print("발송할 템플릿이 없습니다.")
            if hasattr(self, 'status_text'):
                self.status_text.append("❌ 발송할 템플릿이 없습니다. 템플릿 관리에서 템플릿을 먼저 생성해주세요.")
            return
            
        # 이미 발송한 상담 No 목록 확인
        sent_nos = self.sms_tab._get_sent_consultation_nos()
        content = self.sms_tab.templates[0].get("content", "")
        
        # 테스트 모드 확인
        test_mode = self.cfg.getboolean("settings", "test_mode", fallback=True)
        test_receiver = self.cfg.get("settings", "test_receiver", fallback="01012345678")
        
        new_count = 0
        for consultation in consultations:
            consultation_no = consultation["no"]
            phone = consultation["phone"]
            
            # 이미 발송한 건은 스킵
            if consultation_no in sent_nos:
                if hasattr(self, 'status_text'):
                    self.status_text.append(f"⏭️ No.{consultation_no} - 이미 발송완료 (스킵)")
                continue
            
            # 전화번호 유효성 검사
            if not phone or len(phone) < 10:
                if hasattr(self, 'status_text'):
                    self.status_text.append(f"⚠️ No.{consultation_no} - 유효하지 않은 전화번호: '{phone}' (스킵)")
                continue
                
            # 🔥 테스트 모드 처리: 무조건 테스트 번호로 발송
            actual_receiver = test_receiver if test_mode else phone
            
            # SMS 발송
            sms_success = send_sms(actual_receiver, content, self.cfg)
            
            if sms_success:
                # 로그 저장 (상담 No 포함) - 성공한 경우에만
                self.sms_tab._log_sms(actual_receiver, content, consultation_no)
                new_count += 1
                
                # 상태 업데이트
                if hasattr(self, 'status_text'):
                    if test_mode:
                        self.status_text.append(f"🧪 [테스트모드] No.{consultation_no}: {phone} → {actual_receiver} 발송완료")
                    else:
                        self.status_text.append(f"🚀 [실제발송] No.{consultation_no}: {actual_receiver} 발송완료")
            else:
                # SMS 발송 실패
                if hasattr(self, 'status_text'):
                    self.status_text.append(f"❌ No.{consultation_no}: {actual_receiver} 발송실패")
        
        if new_count > 0 and hasattr(self, 'status_text'):
            mode_text = "테스트모드" if test_mode else "실제발송"
            self.status_text.append(f"📊 총 {new_count}건의 신규 상담 문자 발송 완료! ({mode_text})")
        elif len(consultations) > 0 and new_count == 0 and hasattr(self, 'status_text'):
            self.status_text.append(f"ℹ️ {len(consultations)}건의 신규 상담이 감지되었지만 모두 스킵되었습니다.")
    
    # 기존 함수 호환성을 위해 유지 (단일 건 처리)
    def _on_new_phone(self, phone):
        """단일 상담 처리 (기존 호환성)"""
        consultations = [{"no": "", "phone": phone}]
        self._on_new_consultations(consultations)

    def _show_startup_info(self):
        """프로그램 시작 시 초기 상태 정보 표시"""
        if hasattr(self, 'status_text'):
            test_mode = self.cfg.getboolean("settings", "test_mode", fallback=True)
            test_receiver = self.cfg.get("settings", "test_receiver", fallback="01012345678")
            template_count = len(self.sms_tab.templates)
            
            self.status_text.append("=" * 50)
            self.status_text.append("🚀 사이트 감지 & 문자발송 프로그램 시작")
            self.status_text.append("=" * 50)
            self.status_text.append(f"📋 템플릿 설정: {template_count}개 템플릿 로드됨")
            
            if test_mode:
                self.status_text.append(f"🧪 테스트 모드: 활성화 (모든 문자 → {test_receiver})")
            else:
                self.status_text.append("🚀 실제 발송 모드: 활성화 (실제 번호로 발송)")
            
            self.status_text.append("📍 크롤러가 첫 번째 행을 기준점으로 설정 중...")
            self.status_text.append("⏳ 기준점 설정 후 신규 상담 감지 시작됩니다")
            self.status_text.append("=" * 50)

    def _update_history_info(self):
        """발송 이력 정보 업데이트"""
        try:
            sent_nos = self.sms_tab._get_sent_consultation_nos()
            
            history_info = f"""📊 SMS 발송 이력 현황:
• 총 발송 완료 상담: {len(sent_nos)}건
• 시스템 모드: {"🧪 테스트 모드" if self.cfg.getboolean("settings", "test_mode", fallback=True) else "🚀 실제 발송 모드"}
• 테스트 수신번호: {self.cfg.get("settings", "test_receiver", fallback="설정 없음")}

💡 첫 번째 행 기준점으로 신규 상담만 감지합니다."""
            
            self.history_info_text.setText(history_info)
            
        except Exception as e:
            self.history_info_text.setText(f"발송 이력 로드 오류: {e}")


def main():
    app = QApplication(sys.argv)
    
    # 강제 라이트모드 설정
    light_style = """
    QWidget {
        background-color: white;
        color: black;
    }
    QMainWindow {
        background-color: #f0f0f0;
    }
    QTabWidget::pane {
        background-color: white;
        border: 1px solid #c0c0c0;
    }
    QTabBar::tab {
        background-color: #e0e0e0;
        color: black;
        border: 1px solid #c0c0c0;
        padding: 8px 16px;
        margin-right: 2px;
    }
    QTabBar::tab:selected {
        background-color: white;
        border-bottom: 1px solid white;
    }
    QPushButton {
        background-color: #e0e0e0;
        color: black;
        border: 1px solid #c0c0c0;
        padding: 6px 12px;
        border-radius: 4px;
    }
    QPushButton:hover {
        background-color: #d0d0d0;
    }
    QPushButton:pressed {
        background-color: #c0c0c0;
    }
    QLineEdit {
        background-color: white;
        color: black;
        border: 1px solid #c0c0c0;
        padding: 4px;
        border-radius: 2px;
    }
    QTextEdit {
        background-color: white;
        color: black;
        border: 1px solid #c0c0c0;
    }
    QComboBox {
        background-color: white;
        color: black;
        border: 1px solid #c0c0c0;
        padding: 4px;
    }
    QComboBox::drop-down {
        background-color: #e0e0e0;
        border-left: 1px solid #c0c0c0;
    }
    QGroupBox {
        color: black;
        border: 1px solid #c0c0c0;
        border-radius: 4px;
        margin-top: 10px;
        padding-top: 10px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        left: 10px;
        padding: 0 4px 0 4px;
        background-color: white;
    }
    QLabel {
        color: black;
        background-color: transparent;
    }
    QCheckBox {
        color: black;
        background-color: transparent;
        font-weight: bold;
        padding: 4px;
    }
    QCheckBox::indicator {
        width: 18px;
        height: 18px;
        border: 2px solid #666;
        border-radius: 4px;
        background-color: white;
    }
    QCheckBox::indicator:hover {
        border: 2px solid #007acc;
        background-color: #f0f8ff;
    }
    QCheckBox::indicator:checked {
        background-color: #007acc;
        border: 2px solid #007acc;
        image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iMTIiIHZpZXdCb3g9IjAgMCAxMiAxMiIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj4KPHBhdGggZD0iTTEwIDNMNC4yNSA4Ljc1TDIgNi41IiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCIvPgo8L3N2Zz4K);
    }
    QCheckBox::indicator:checked:hover {
        background-color: #0056b3;
        border: 2px solid #0056b3;
    }
    """
    app.setStyleSheet(light_style)
    
    mw = MainWindow()
    mw.resize(800, 600)
    mw.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
